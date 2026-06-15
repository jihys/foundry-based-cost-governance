"""Tests for src.auth.key_export — Terraform Output JSON → Excel key sheet."""
from pathlib import Path

import openpyxl
import pytest

from src.auth.key_export import generate_key_excel

EXPECTED_COLUMNS = [
    "Team",
    "Subscription ID",
    "Resource Group",
    "Region",
    "Foundry Endpoint",
    "API Key (Key1)",
    "API Key (Key2)",
    "Model Deployments",
]


# ── Fixture helpers ──────────────────────────────────────────────


def _team(subscription_id: str, resource_group: str, regions: dict) -> dict:
    return {
        "subscription_id": subscription_id,
        "resource_group": resource_group,
        "regions": regions,
    }


def _region(endpoint: str, key1: str, key2: str, model_deployments: list | None = None) -> dict:
    return {
        "endpoint": endpoint,
        "key1": key1,
        "key2": key2,
        "model_deployments": model_deployments or [],
    }


# ── Tests ────────────────────────────────────────────────────────


class TestGenerateKeyExcel:
    """Behavior tests for generate_key_excel through public interface."""

    def test_single_team_single_region(self, tmp_path: Path) -> None:
        """One team with one region produces exactly one data row with correct values."""
        tf = {
            "catalog": _team(
                "sub-cat-001",
                "rg-catalog",
                {
                    "koreacentral": _region(
                        "https://ai-catalog-krc.cognitiveservices.azure.com",
                        "cat-key1-krc",
                        "cat-key2-krc",
                        ["gpt-4o"],
                    ),
                },
            ),
        }
        out = tmp_path / "keys.xlsx"

        generate_key_excel(tf, str(out))

        wb = openpyxl.load_workbook(out)
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 1
        row = rows[0]
        assert row[0] == "catalog"
        assert row[1] == "sub-cat-001"
        assert row[2] == "rg-catalog"
        assert row[3] == "koreacentral"
        assert row[4] == "https://ai-catalog-krc.cognitiveservices.azure.com"
        assert row[5] == "cat-key1-krc"
        assert row[6] == "cat-key2-krc"
        assert row[7] == "gpt-4o"

    def test_multiple_teams_multiple_regions(self, tmp_path: Path) -> None:
        """3 teams (catalog 2 regions, image 1, search 1) → 4 data rows."""
        tf = {
            "catalog": _team("sub-cat", "rg-catalog", {
                "koreacentral": _region("https://cat-krc", "ck1k", "ck2k", ["gpt-4o"]),
                "swedencentral": _region("https://cat-sdc", "ck1s", "ck2s", ["o3-mini"]),
            }),
            "image": _team("sub-img", "rg-image", {
                "koreacentral": _region("https://img-krc", "ik1k", "ik2k", ["dall-e-3"]),
            }),
            "search": _team("sub-srch", "rg-search", {
                "koreacentral": _region("https://srch-krc", "sk1k", "sk2k", ["text-embedding-3-large"]),
            }),
        }
        out = tmp_path / "keys.xlsx"

        generate_key_excel(tf, str(out))

        wb = openpyxl.load_workbook(out)
        rows = list(wb.active.iter_rows(min_row=2, values_only=True))
        assert len(rows) == 4

        team_names = [r[0] for r in rows]
        assert team_names.count("catalog") == 2
        assert team_names.count("image") == 1
        assert team_names.count("search") == 1

    def test_column_schema(self, tmp_path: Path) -> None:
        """Header row contains exactly the 8 expected columns in order."""
        tf = {"t": _team("s", "r", {"kr": _region("e", "k1", "k2")})}
        out = tmp_path / "keys.xlsx"

        generate_key_excel(tf, str(out))

        wb = openpyxl.load_workbook(out)
        headers = [cell.value for cell in wb.active[1]]
        assert headers == EXPECTED_COLUMNS

    def test_model_deployments_comma_separated(self, tmp_path: Path) -> None:
        """Multiple model deployments are joined with ', ' in the cell."""
        tf = {
            "catalog": _team("s", "rg", {
                "koreacentral": _region("e", "k1", "k2", ["gpt-4o", "text-embedding-3-large", "o3-mini"]),
            }),
        }
        out = tmp_path / "keys.xlsx"

        generate_key_excel(tf, str(out))

        wb = openpyxl.load_workbook(out)
        model_cell = wb.active.cell(row=2, column=8).value
        assert model_cell == "gpt-4o, text-embedding-3-large, o3-mini"

    def test_empty_teams_map(self, tmp_path: Path) -> None:
        """Empty terraform_outputs dict → Excel with headers only, zero data rows."""
        out = tmp_path / "keys.xlsx"

        generate_key_excel({}, str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert headers == EXPECTED_COLUMNS
        assert len(data_rows) == 0

    def test_output_path_returns_path_and_file_exists(self, tmp_path: Path) -> None:
        """Function returns a Path object pointing to the created file."""
        tf = {"t": _team("s", "r", {"kr": _region("e", "k1", "k2")})}
        out = tmp_path / "keys.xlsx"

        result = generate_key_excel(tf, str(out))

        assert isinstance(result, Path)
        assert result.exists()
        assert result == out

    def test_append_mode(self, tmp_path: Path) -> None:
        """When the output file already exists, new rows are appended (not overwritten)."""
        out = tmp_path / "keys.xlsx"

        # First call: catalog team
        tf1 = {"catalog": _team("sub-cat", "rg-cat", {
            "koreacentral": _region("https://cat-krc", "ck1", "ck2", ["gpt-4o"]),
        })}
        generate_key_excel(tf1, str(out))

        # Second call: image team (should append)
        tf2 = {"image": _team("sub-img", "rg-img", {
            "koreacentral": _region("https://img-krc", "ik1", "ik2", ["dall-e-3"]),
        })}
        generate_key_excel(tf2, str(out))

        wb = openpyxl.load_workbook(out)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))

        assert headers == EXPECTED_COLUMNS
        assert len(data_rows) == 2
        assert data_rows[0][0] == "catalog"
        assert data_rows[1][0] == "image"
